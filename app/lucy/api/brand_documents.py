from __future__ import annotations

import asyncio
import csv
import io
import json
import re
from collections import defaultdict
from typing import Any, Dict, Iterable, List, Literal, Optional, Tuple, Type

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from loguru import logger
from openai import OpenAI
from openpyxl import load_workbook
from pydantic import BaseModel, HttpUrl, ValidationError, field_validator

from lucy.agents.clients.openai_client import get_openai_client
from lucy.database.supabase_client import get_client, get_user_org_profiles
from lucy.utils.auth import extract_user_id, verify_jwt

router = APIRouter()

DocType = Literal["products-doc", "competitors-doc", "audiences-doc", "locations-doc"]

ALLOWED_EXTENSIONS = {"csv", "xlsx"}
MIME_TO_EXTENSION = {
    "text/csv": "csv",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    "application/vnd.ms-excel": "xls",
}

# ------------------------------------------------------------
# Limits
# ------------------------------------------------------------
MAX_UPLOAD_BYTES = 50 * 1024 * 1024
MAX_XLSX_SHEETS = 10
MAX_TOTAL_ROWS = 5000
MAX_TOTAL_COLUMNS = 100
MAX_FINALIZE_CANDIDATES = 1500
MAX_LLM_INPUT_CHARS = 180_000

MODEL_NAME = "gpt-5-mini"

# ------------------------------------------------------------
# Response models
# ------------------------------------------------------------
class ProductItem(BaseModel):
    name: str
    description: str


class CompetitorItem(BaseModel):
    name: str
    url: HttpUrl
    type: Literal["direct", "indirect", "replacement"]


class AudienceItem(BaseModel):
    name: str
    description: str


class AddressPart(BaseModel):
    part_name: str
    index_offset: int


class AddressObj(BaseModel):
    place_name: str
    place_name_parts: List[AddressPart]


class HoursObj(BaseModel):
    day_selection: Literal["all", "weekdays", "weekends", "mon", "tue", "wed", "thu", "fri", "sat", "sun"]
    start_time: str
    end_time: str

    @field_validator("start_time", "end_time")
    @classmethod
    def validate_hhmm(cls, v: str) -> str:
        if len(v) != 5 or v[2] != ":":
            raise ValueError("Time must be HH:mm")
        hh, mm = v.split(":")
        if not (hh.isdigit() and mm.isdigit()):
            raise ValueError("Time must be HH:mm")
        if not (0 <= int(hh) <= 23 and 0 <= int(mm) <= 59):
            raise ValueError("Time must be HH:mm")
        return v


class TargetingObj(BaseModel):
    type: Literal["radius", "dma", "zip"]
    value: str
    unit: Optional[str] = None


class LocationItem(BaseModel):
    name: str
    address: AddressObj
    timezone: Optional[str] = None
    tags: Optional[List[str]] = None
    targeting: Optional[TargetingObj] = None
    phone: Optional[str] = None
    landingPage: Optional[str] = None
    hours: Optional[List[HoursObj]] = None


class ProductResponse(BaseModel):
    analysisResults: List[ProductItem]


class CompetitorResponse(BaseModel):
    analysisResults: List[CompetitorItem]


class AudienceResponse(BaseModel):
    analysisResults: List[AudienceItem]


class LocationResponse(BaseModel):
    analysisResults: List[LocationItem]


RESPONSE_MODEL_BY_DOC_TYPE: Dict[str, Type[BaseModel]] = {
    "products-doc": ProductResponse,
    "competitors-doc": CompetitorResponse,
    "audiences-doc": AudienceResponse,
    "locations-doc": LocationResponse,
}

# ------------------------------------------------------------
# Intermediate models
# ------------------------------------------------------------
class CandidateSource(BaseModel):
    unused_columns: Dict[str, str]


class ParsedSheet(BaseModel):
    title: str
    columns: List[str]
    rows: List[Dict[str, str]]


class ParsedTabularFile(BaseModel):
    kind: Literal["tabular"] = "tabular"
    sheets: List[ParsedSheet]


class ProductCandidate(BaseModel):
    name: str
    description_source: Optional[str] = None
    source: CandidateSource


class AudienceCandidate(BaseModel):
    name: str
    description_source: Optional[str] = None
    source: CandidateSource


class CompetitorCandidate(BaseModel):
    name: str
    url: Optional[str] = None
    type_hint: Optional[str] = None
    source: CandidateSource


class LocationCandidate(BaseModel):
    name: str
    address_raw: Optional[str] = None
    phone: Optional[str] = None
    landingPage: Optional[str] = None
    timezone: Optional[str] = None
    tags: Optional[List[str]] = None
    hours_raw: Optional[str] = None
    targeting_raw: Optional[str] = None
    source: CandidateSource


# ------------------------------------------------------------
# Auth helpers
# ------------------------------------------------------------
def _authorized_org_ids(user_id: str) -> set[str]:
    rows = get_user_org_profiles(user_id) or []
    return {r["org_id"] for r in rows if isinstance(r, dict) and r.get("org_id")}


async def _get_verified_user(payload: Dict[str, Any]) -> str:
    user_id = extract_user_id(payload)
    if not user_id or user_id == "anonymous":
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Missing user identity",
        )
    return user_id


def _verify_brand_access(user_id: str, brand_id: str) -> None:
    sb = get_client()
    try:
        result = (
            sb.table("brands")
            .select("brand_id, associated_organization_id")
            .eq("brand_id", brand_id)
            .limit(1)
            .execute()
        )
        rows = getattr(result, "data", None) or []
        if not rows:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Brand not found")

        brand = rows[0]
        org_id = brand.get("associated_organization_id")
        if not org_id or org_id not in _authorized_org_ids(user_id):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized for brand_id")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Brand access verification failed: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to verify brand access",
        )


# ------------------------------------------------------------
# Brand config / prompts
# ------------------------------------------------------------
def get_brand_config(brand_id: str) -> Optional[Dict[str, Any]]:
    sb = get_client()
    result = (
        sb.table("brands")
        .select(
            """
            brand_name,
            brand_website_url,
            brand_core_values,
            brand_keyword_blacklist,
            brand_tone_of_voice,
            brand_mission_vision,
            brand_positioning,
            brand_tagline
            """
        )
        .eq("brand_id", brand_id)
        .limit(1)
        .execute()
    )
    rows = getattr(result, "data", None) or []
    return rows[0] if rows else None


def get_brand_prompt(brand_config: Optional[Dict[str, Any]]) -> str:
    if not brand_config:
        return "BRAND PROFILE:\nNo brand profile available."

    parts = ["BRAND PROFILE:"]
    if brand_config.get("brand_name") and brand_config["brand_name"] not in {"New Primary Brand", "New Organization"}:
        parts.append(f"Brand name: {brand_config['brand_name']}")
    if brand_config.get("brand_website_url") and brand_config["brand_website_url"] != "https://example.com":
        parts.append(f"Brand website: {brand_config['brand_website_url']}")
    if brand_config.get("brand_keyword_blacklist"):
        parts.append(f"Brand blacklisted keywords: {';'.join(brand_config['brand_keyword_blacklist'])}")
    if brand_config.get("brand_core_values"):
        parts.append(f"Brand core values: {brand_config['brand_core_values']}")
    if brand_config.get("brand_tone_of_voice"):
        parts.append(f"Brand tone of voice: {brand_config['brand_tone_of_voice']}")
    if brand_config.get("brand_mission_vision"):
        parts.append(f"Brand mission/vision: {brand_config['brand_mission_vision']}")
    if brand_config.get("brand_positioning"):
        parts.append(f"Brand positioning: {brand_config['brand_positioning']}")
    if brand_config.get("brand_tagline"):
        parts.append(f"Brand tagline: {brand_config['brand_tagline']}")
    return "\n".join(parts)


def get_finalize_system_prompt(doc_type: DocType) -> str:
    brand_audiences_system_prompt = """
You are the Lofi Marketing Agent (Lucy). Your sole responsibility is to generate accurate target audience segments for the brand.

You MUST analyze and use, in priority order:
1. Candidate records extracted from the uploaded tabular file
2. The candidate _source.unused_columns object
3. The brand profile

Candidate fields are the parser's primary extraction and must be treated as the main source of truth.
You may look at _source.unused_columns ONLY to enrich or refine fields that already exist on the candidate.
You must NOT use _source.unused_columns to invent brand new entities.
You must NOT replace a clearly supported candidate field with a weaker guess from unused_columns.
Use unused_columns conservatively to improve descriptions, clarify meaning, or fill already-existing optional detail when strongly supported.

Obligatory general rules:
- Return JSON strictly matching the provided schema
- Segments must be distinct and actionable; avoid generic or filler names
- Each segment description must include:
  - Who they are
  - Key needs or pain points
  - Why the brand fits
- Be explicit when inferring from incomplete data
- JSON only
""".strip()

    brand_products_system_prompt = """
You are the Lofi Marketing Agent (Lucy). Your sole responsibility is to generate a list of the brand's products or product lines.

You MUST analyze and use, in priority order:
1. Candidate records extracted from the uploaded tabular file
2. The candidate _source.unused_columns object
3. The brand profile

Candidate fields are the parser's primary extraction and must be treated as the main source of truth.
You may look at _source.unused_columns ONLY to enrich or refine fields that already exist on the candidate.
You must NOT use _source.unused_columns to invent brand new products or product lines.
You must NOT replace a clearly supported candidate field with a weaker guess from unused_columns.
Use unused_columns conservatively to improve descriptions, clarify use cases, audience fit, benefits, or differentiators when strongly supported.

Obligatory general rules:
- Return JSON strictly matching the provided schema
- Each product or product line must be distinct and clearly named
- Each description must explain:
  - What it is
  - Who it is for
  - Key value or benefits
  - Main differentiator
- Keep descriptions specific, accurate, and marketing-useful
- JSON only
""".strip()

    brand_competitors_system_prompt = """
You are the Lofi Marketing Agent (Lucy). Your sole responsibility is to generate a list of the brand's competitors.

You MUST analyze and use, in priority order:
1. Candidate records extracted from the uploaded tabular file
2. The candidate _source.unused_columns object
3. The brand profile

Candidate fields are the parser's primary extraction and must be treated as the main source of truth.
You may look at _source.unused_columns ONLY to enrich or refine fields that already exist on the candidate.
You must NOT use _source.unused_columns to invent brand new competitors.
You must NOT replace a clearly supported candidate field with a weaker guess from unused_columns.
Use unused_columns conservatively to clarify competitor relationship or validate homepage selection when strongly supported.

Obligatory general rules:
- Return JSON strictly matching the provided schema
- Each competitor must be a real company or brand as supported by the provided data
- Include ONLY the competitor's official homepage URL
- Do not invent unsupported URLs
- The "type" field must clearly describe the competitive relationship and MUST be exactly one of:
  'direct', 'indirect', 'replacement'
- Keep competitors tightly relevant to the brand's market and positioning
- JSON only
""".strip()

    brand_locations_system_prompt = """
You are the Lofi Marketing Agent (Lucy). Your sole responsibility is to generate a list of the brand's real physical locations.

You MUST analyze and use, in priority order:
1. Candidate records extracted from the uploaded tabular file
2. The candidate _source.unused_columns object
3. The brand profile

Candidate fields are the parser's primary extraction and must be treated as the main source of truth.
You may look at _source.unused_columns ONLY to enrich or refine fields that already exist on the candidate.
You must NOT use _source.unused_columns to invent brand new locations.
You must NOT replace a clearly supported candidate field with a weaker guess from unused_columns.

A valid location MUST:
- Be clearly associated with the brand
- Represent a physical, visitable place
- Contain a clear, identifiable address

You MUST exclude:
- Entries without a clear physical address
- Service areas / coverage regions
- P.O. boxes only
- Third-party retailers / partners / resellers
- Permanently closed / relocated locations

OUTPUT RULES:
- Return valid JSON strictly matching the schema
- Populate optional fields ONLY if explicitly supported
- Per location:
  - ADDRESS (AutocompletePrediction-like; REQUIRED):
    - address.place_name must be a Google-style place description string
    - address.place_name_parts must contain exact substrings of place_name with correct 0-based index_offset values
  - NAME (REQUIRED):
    - Use the strongest supported location/store/office/branch/clinic/facility/showroom/dealer/distributor name from the candidate
  - TAGS (OPTIONAL):
    - First tag MUST be connected to address.place_name in a format: "City, State" (e.g., "Auburn, AL", "Charlotte, NC")
    - The rest of tags should be unique info about specific location, include ONLY if explicitly listed
  - HOURS (OPTIONAL):
    - Include only if explicitly supported
    - start_time / end_time must be HH:mm (24h)
  - PHONE (OPTIONAL):
    - Include only if explicitly supported in E.164 format
FINAL RULES:
- NO duplicates
- NO explanations
- NO invention
- JSON ONLY
""".strip()

    if doc_type == "products-doc":
        return brand_products_system_prompt
    if doc_type == "competitors-doc":
        return brand_competitors_system_prompt
    if doc_type == "locations-doc":
        return brand_locations_system_prompt
    return brand_audiences_system_prompt


def build_finalize_user_prompt(
    doc_type: DocType,
    brand_prompt: str,
    candidates_payload: Dict[str, Any],
) -> str:
    return f"""
Analyze these pre-extracted tabular candidates for `{doc_type}`.

{brand_prompt}

Important instructions for this task:
- Candidate top-level fields are the primary parsed extraction and should be preserved unless there is strong evidence to refine them.
- Each candidate may include `_source.unused_columns`.
- You may inspect `_source.unused_columns` ONLY to enrich or refine fields that already exist on that candidate.
- Do NOT create new entities from `_source.unused_columns`.
- Do NOT use `_source.unused_columns` to override strong candidate fields with weaker guesses.
- Use `_source.unused_columns` only when it clearly improves specificity, completeness, or formatting of already existing data.

CANDIDATES:
{json.dumps(candidates_payload, ensure_ascii=False, separators=(",", ":"))}

Return JSON only.
""".strip()


def get_schema(doc_type: DocType) -> Dict[str, Any]:
    if doc_type == "competitors-doc":
        return {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "analysisResults": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "additionalProperties": False,
                        "properties": {
                            "name": {"type": "string"},
                            "url": {"type": "string"},
                            "type": {"type": "string", "enum": ["direct", "indirect", "replacement"]},
                        },
                        "required": ["name", "url", "type"],
                    },
                }
            },
            "required": ["analysisResults"],
        }

    if doc_type == "products-doc":
        return {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "analysisResults": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "additionalProperties": False,
                        "properties": {
                            "name": {"type": "string"},
                            "description": {"type": "string"},
                        },
                        "required": ["name", "description"],
                    },
                }
            },
            "required": ["analysisResults"],
        }

    if doc_type == "locations-doc":
        return {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "analysisResults": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "additionalProperties": False,
                        "properties": {
                            "name": {"type": "string"},
                            "address": {
                                "type": "object",
                                "additionalProperties": False,
                                "properties": {
                                    "place_name": {"type": "string"},
                                    "place_name_parts": {
                                        "type": "array",
                                        "items": {
                                            "type": "object",
                                            "additionalProperties": False,
                                            "properties": {
                                                "part_name": {"type": "string"},
                                                "index_offset": {"type": "integer"},
                                            },
                                            "required": ["part_name", "index_offset"],
                                        },
                                    },
                                },
                                "required": ["place_name", "place_name_parts"],
                            },
                            "timezone": {"type": ["string", "null"]},
                            "tags": {"type": ["array", "null"], "items": {"type": "string"}},
                            "targeting": {
                                "type": ["object", "null"],
                                "additionalProperties": False,
                                "properties": {
                                    "type": {"type": "string", "enum": ["radius", "dma", "zip"]},
                                    "value": {"type": "string"},
                                    "unit": {"type": ["string", "null"]},
                                },
                                "required": ["type", "value", "unit"],
                            },
                            "phone": {"type": ["string", "null"]},
                            "landingPage": {"type": ["string", "null"]},
                            "hours": {
                                "type": ["array", "null"],
                                "items": {
                                    "type": "object",
                                    "additionalProperties": False,
                                    "properties": {
                                        "day_selection": {
                                            "type": "string",
                                            "enum": ["all", "weekdays", "weekends", "mon", "tue", "wed", "thu", "fri", "sat", "sun"],
                                        },
                                        "start_time": {"type": "string"},
                                        "end_time": {"type": "string"},
                                    },
                                    "required": ["day_selection", "start_time", "end_time"],
                                },
                            },
                        },
                        "required": ["name", "address"],
                    },
                }
            },
            "required": ["analysisResults"],
        }

    return {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "analysisResults": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "name": {"type": "string"},
                        "description": {"type": "string"},
                    },
                    "required": ["name", "description"],
                },
            }
        },
        "required": ["analysisResults"],
    }


# ------------------------------------------------------------
# Generic helpers
# ------------------------------------------------------------
def detect_extension(filename: str, mime_type: Optional[str]) -> str:
    filename = (filename or "").lower()
    if "." in filename:
        ext = filename.rsplit(".", 1)[1]
        if ext in ALLOWED_EXTENSIONS:
            return ext

    if mime_type:
        ext = MIME_TO_EXTENSION.get(mime_type.lower())
        if ext == "xls":
            raise HTTPException(status_code=400, detail="Only csv and xlsx are allowed for this endpoint")
        if ext in ALLOWED_EXTENSIONS:
            return ext

    raise HTTPException(status_code=400, detail="Only csv and xlsx are allowed for this endpoint")


def _clean_text(v: Any) -> str:
    if v is None:
        return ""
    text = str(v).replace("\x00", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_header(v: str) -> str:
    s = _clean_text(v).lower()
    s = s.replace("&", " and ")
    s = re.sub(r"[\W_]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _normalize_key(v: str) -> str:
    s = _clean_text(v).lower()
    s = re.sub(r"\s+", " ", s)
    return s.strip(" -_|/.,;")


def _normalize_url(url: str) -> str:
    url = _clean_text(url)
    if not url:
        return ""
    if not re.match(r"^https?://", url, flags=re.I):
        url = "https://" + url
    return url.rstrip("/")


def _normalize_phone(phone: str) -> str:
    raw = _clean_text(phone)
    if not raw:
        return ""
    digits = re.sub(r"[^\d+]", "", raw)
    return digits or raw


def _is_blank_row(row: Dict[str, str]) -> bool:
    return not any(_clean_text(v) for v in row.values())


def _compact_text(*parts: Optional[str], max_len: int = 1000) -> str:
    joined = " | ".join([_clean_text(p) for p in parts if _clean_text(p)])
    if len(joined) > max_len:
        return joined[:max_len].rstrip() + "..."
    return joined


def _find_part_offsets(place_name: str) -> List[AddressPart]:
    parts: List[AddressPart] = []
    seen: set[tuple[str, int]] = set()

    for piece in [p.strip() for p in place_name.split(",") if p.strip()]:
        idx = place_name.find(piece)
        if idx >= 0 and (piece, idx) not in seen:
            parts.append(AddressPart(part_name=piece, index_offset=idx))
            seen.add((piece, idx))

    if not parts and place_name:
        parts.append(AddressPart(part_name=place_name, index_offset=0))

    return parts


def _build_candidate_source(
    *,
    row: Dict[str, str],
    used_column_names: Iterable[Optional[str]],
) -> CandidateSource:
    used_set = {c for c in used_column_names if c}
    unused_columns = {k: v for k, v in row.items() if k not in used_set and _clean_text(v)}
    return CandidateSource(unused_columns=unused_columns)


# ------------------------------------------------------------
# Tabular parsing
# ------------------------------------------------------------
def _read_csv_rows(file_bytes: bytes, max_rows: int, max_cols: int) -> ParsedSheet:
    encodings = ["utf-8-sig", "utf-8", "latin-1"]
    last_error: Optional[Exception] = None

    for enc in encodings:
        try:
            text = file_bytes.decode(enc, errors="replace")
            sample = text[:8192]
            try:
                dialect = csv.Sniffer().sniff(sample)
            except Exception:
                dialect = csv.excel

            reader = csv.reader(io.StringIO(text), dialect=dialect)
            raw_rows: List[List[str]] = []

            for i, row in enumerate(reader):
                if i >= max_rows + 20:
                    break
                raw_rows.append([_clean_text(c) for c in row])

            if not raw_rows:
                raise ValueError("CSV appears empty")

            header_idx, header = _choose_header_row(raw_rows)
            data_rows = raw_rows[header_idx + 1 :]

            if len(header) > max_cols:
                raise ValueError(f"CSV has too many columns ({len(header)} > {max_cols})")

            rows = _rows_to_dicts(header, data_rows, max_rows=max_rows)
            return ParsedSheet(title="CSV DATA", columns=header, rows=rows)

        except Exception as e:
            last_error = e

    raise ValueError(f"Failed to parse CSV: {last_error}")


def _read_xlsx_rows(file_bytes: bytes, max_sheets: int, max_rows: int, max_cols: int) -> List[ParsedSheet]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_names = wb.sheetnames[:max_sheets]

    if len(wb.sheetnames) > max_sheets:
        raise ValueError(f"Too many sheets ({len(wb.sheetnames)} > {max_sheets})")

    parsed_sheets: List[ParsedSheet] = []
    total_rows_seen = 0

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        raw_rows: List[List[str]] = []

        for row in ws.iter_rows(values_only=True):
            cleaned = [_clean_text(c) for c in row]
            raw_rows.append(cleaned)
            if len(raw_rows) >= max_rows + 20:
                break

        if not raw_rows:
            parsed_sheets.append(ParsedSheet(title=sheet_name, columns=[], rows=[]))
            continue

        header_idx, header = _choose_header_row(raw_rows)

        if len(header) > max_cols:
            raise ValueError(f"Sheet '{sheet_name}' has too many columns ({len(header)} > {max_cols})")

        data_rows = raw_rows[header_idx + 1 :]
        rows = _rows_to_dicts(header, data_rows, max_rows=max_rows)

        total_rows_seen += len(rows)
        if total_rows_seen > MAX_TOTAL_ROWS:
            raise ValueError(f"XLSX has too many rows ({total_rows_seen} > {MAX_TOTAL_ROWS})")

        parsed_sheets.append(ParsedSheet(title=sheet_name, columns=header, rows=rows))

    return parsed_sheets


def _choose_header_row(raw_rows: List[List[str]]) -> Tuple[int, List[str]]:
    search_rows = raw_rows[:10]
    best_idx = 0
    best_score = float("-inf")
    best_header: List[str] = []

    for idx, row in enumerate(search_rows):
        cells = [_clean_text(c) for c in row]
        non_empty = [c for c in cells if c]
        if not non_empty:
            continue

        unique_ratio = len(set(c.lower() for c in non_empty)) / max(1, len(non_empty))
        avg_len = sum(len(c) for c in non_empty) / max(1, len(non_empty))
        alpha_ratio = sum(1 for c in non_empty if re.search(r"[A-Za-z]", c)) / max(1, len(non_empty))

        score = (len(non_empty) * 2.0) + (unique_ratio * 3.0) + (alpha_ratio * 2.0) - (avg_len / 40.0)

        if score > best_score:
            best_score = score
            best_idx = idx
            best_header = cells

    header = [_clean_text(c) or f"column_{i+1}" for i, c in enumerate(best_header)]
    header = _dedupe_headers(header)
    return best_idx, header


def _dedupe_headers(headers: List[str]) -> List[str]:
    seen: Dict[str, int] = defaultdict(int)
    output: List[str] = []

    for h in headers:
        base = _clean_text(h) or "column"
        seen[base] += 1
        if seen[base] == 1:
            output.append(base)
        else:
            output.append(f"{base}_{seen[base]}")
    return output


def _rows_to_dicts(header: List[str], rows: List[List[str]], max_rows: int) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    width = len(header)

    for row in rows:
        row = row[:width] + [""] * max(0, width - len(row))
        item = {header[i]: _clean_text(row[i]) for i in range(width)}
        if _is_blank_row(item):
            continue
        out.append(item)
        if len(out) >= max_rows:
            break

    return out


def parse_tabular_file(file_bytes: bytes, extension: str) -> ParsedTabularFile:
    if extension == "csv":
        sheet = _read_csv_rows(file_bytes, max_rows=MAX_TOTAL_ROWS, max_cols=MAX_TOTAL_COLUMNS)
        return ParsedTabularFile(sheets=[sheet])

    if extension == "xlsx":
        sheets = _read_xlsx_rows(
            file_bytes,
            max_sheets=MAX_XLSX_SHEETS,
            max_rows=min(MAX_TOTAL_ROWS, 2000),
            max_cols=MAX_TOTAL_COLUMNS,
        )
        return ParsedTabularFile(sheets=sheets)

    raise ValueError(f"Unsupported extension: {extension}")


# ------------------------------------------------------------
# Column meaning inference
# ------------------------------------------------------------
COLUMN_HINTS: Dict[str, List[str]] = {
    "name": ["name", "product", "product name", "store", "store name", "location", "branch", "company", "brand"],
    "description": ["description", "details", "about", "summary", "overview", "benefits", "notes"],
    "url": ["url", "website", "web", "homepage", "site", "landing page", "link"],
    "phone": ["phone", "telephone", "tel", "mobile", "contact number"],
    "timezone": ["timezone", "time zone", "tz"],
    "hours": ["hours", "opening hours", "business hours", "store hours", "schedule"],
    "address": ["address", "street", "street address", "address line 1", "address1", "addr"],
    "address2": ["address line 2", "address2", "suite", "unit"],
    "city": ["city", "town"],
    "state": ["state", "province", "region"],
    "zip": ["zip", "postal", "postal code", "zipcode", "zip code"],
    "country": ["country"],
    "tags": ["tags", "labels", "categories", "category"],
    "targeting": ["targeting", "radius", "dma", "market", "coverage", "zip targeting"],
    "type": ["type", "competitor type", "kind"],
    "audience": ["audience", "segment", "persona", "target audience"],
}


def _infer_column_roles(columns: List[str]) -> Dict[str, str]:
    normalized = {col: _normalize_header(col) for col in columns}
    assigned: Dict[str, str] = {}

    for role, patterns in COLUMN_HINTS.items():
        best_col = None
        best_score = 0

        for original, norm in normalized.items():
            score = 0
            for p in patterns:
                p_norm = _normalize_header(p)
                if norm == p_norm:
                    score = max(score, 100)
                elif p_norm in norm:
                    score = max(score, 80)
                elif any(tok == p_norm for tok in norm.split()):
                    score = max(score, 60)

            if score > best_score:
                best_score = score
                best_col = original

        if best_col:
            assigned[role] = best_col

    return assigned


# ------------------------------------------------------------
# Candidate builders
# ------------------------------------------------------------
def build_candidates(parsed: ParsedTabularFile, doc_type: DocType) -> List[BaseModel]:
    if doc_type == "products-doc":
        return build_product_candidates(parsed)
    if doc_type == "audiences-doc":
        return build_audience_candidates(parsed)
    if doc_type == "competitors-doc":
        return build_competitor_candidates(parsed)
    if doc_type == "locations-doc":
        return build_location_candidates(parsed)
    raise ValueError(f"Unsupported doc_type: {doc_type}")


def build_product_candidates(parsed: ParsedTabularFile) -> List[ProductCandidate]:
    out: List[ProductCandidate] = []

    for sheet in parsed.sheets:
        if not sheet.rows or not sheet.columns:
            continue

        roles = _infer_column_roles(sheet.columns)
        name_col = roles.get("name") or _best_fallback_name_col(sheet.columns)
        desc_col = roles.get("description")

        for row in sheet.rows:
            name = _clean_text(row.get(name_col, "")) if name_col else ""
            if not _looks_like_entity_name(name):
                continue

            description = _clean_text(row.get(desc_col, "")) if desc_col else ""
            if not description:
                description = _compact_text(*row.values(), max_len=800)

            used_cols = [name_col, desc_col]
            source = _build_candidate_source(row=row, used_column_names=used_cols)

            out.append(
                ProductCandidate(
                    name=name,
                    description_source=description or None,
                    source=source,
                )
            )

    return merge_product_candidates(out)


def build_audience_candidates(parsed: ParsedTabularFile) -> List[AudienceCandidate]:
    out: List[AudienceCandidate] = []

    for sheet in parsed.sheets:
        if not sheet.rows or not sheet.columns:
            continue

        roles = _infer_column_roles(sheet.columns)
        name_col = roles.get("audience") or roles.get("name") or _best_fallback_name_col(sheet.columns)
        desc_col = roles.get("description")

        for row in sheet.rows:
            name = _clean_text(row.get(name_col, "")) if name_col else ""
            if not _looks_like_entity_name(name):
                continue

            description = _clean_text(row.get(desc_col, "")) if desc_col else ""
            if not description:
                description = _compact_text(*row.values(), max_len=800)

            used_cols = [name_col, desc_col]
            source = _build_candidate_source(row=row, used_column_names=used_cols)

            out.append(
                AudienceCandidate(
                    name=name,
                    description_source=description or None,
                    source=source,
                )
            )

    return merge_audience_candidates(out)


def build_competitor_candidates(parsed: ParsedTabularFile) -> List[CompetitorCandidate]:
    out: List[CompetitorCandidate] = []

    for sheet in parsed.sheets:
        if not sheet.rows or not sheet.columns:
            continue

        roles = _infer_column_roles(sheet.columns)
        name_col = roles.get("name") or _best_fallback_name_col(sheet.columns)
        url_col = roles.get("url")
        type_col = roles.get("type")

        for row in sheet.rows:
            name = _clean_text(row.get(name_col, "")) if name_col else ""
            if not _looks_like_entity_name(name):
                continue

            url = _clean_text(row.get(url_col, "")) if url_col else ""
            type_hint = _clean_text(row.get(type_col, "")) if type_col else ""

            used_cols = [name_col, url_col, type_col]
            source = _build_candidate_source(row=row, used_column_names=used_cols)

            out.append(
                CompetitorCandidate(
                    name=name,
                    url=_normalize_url(url) if url else None,
                    type_hint=type_hint or None,
                    source=source,
                )
            )

    return merge_competitor_candidates(out)


def build_location_candidates(parsed: ParsedTabularFile) -> List[LocationCandidate]:
    out: List[LocationCandidate] = []

    for sheet in parsed.sheets:
        if not sheet.rows or not sheet.columns:
            continue

        roles = _infer_column_roles(sheet.columns)

        name_col = roles.get("name") or _best_location_name_col(sheet.columns)
        phone_col = roles.get("phone")
        url_col = roles.get("url")
        timezone_col = roles.get("timezone")
        hours_col = roles.get("hours")
        tags_col = roles.get("tags")
        targeting_col = roles.get("targeting")

        address_cols = {
            "address": roles.get("address"),
            "address2": roles.get("address2"),
            "city": roles.get("city"),
            "state": roles.get("state"),
            "zip": roles.get("zip"),
            "country": roles.get("country"),
        }

        used_cols = [
            name_col,
            phone_col,
            url_col,
            timezone_col,
            hours_col,
            tags_col,
            targeting_col,
            address_cols.get("address"),
            address_cols.get("address2"),
            address_cols.get("city"),
            address_cols.get("state"),
            address_cols.get("zip"),
            address_cols.get("country"),
        ]

        for row in sheet.rows:
            name = _clean_text(row.get(name_col, "")) if name_col else ""
            address_raw = _assemble_address_from_row(row, address_cols)

            if not name and not address_raw:
                continue

            if not name and address_raw:
                name = address_raw

            if not _looks_like_location_candidate(name, address_raw):
                continue

            tags = None
            if tags_col:
                raw_tags = _clean_text(row.get(tags_col, ""))
                if raw_tags:
                    tags = [t.strip() for t in re.split(r"[|,;/]", raw_tags) if t.strip()] or None

            source = _build_candidate_source(row=row, used_column_names=used_cols)

            out.append(
                LocationCandidate(
                    name=name,
                    address_raw=address_raw or None,
                    phone=_normalize_phone(row.get(phone_col, "")) if phone_col else None,
                    landingPage=_normalize_url(row.get(url_col, "")) if url_col and row.get(url_col) else None,
                    timezone=_clean_text(row.get(timezone_col, "")) if timezone_col else None,
                    tags=tags,
                    hours_raw=_clean_text(row.get(hours_col, "")) if hours_col else None,
                    targeting_raw=_clean_text(row.get(targeting_col, "")) if targeting_col else None,
                    source=source,
                )
            )

    return merge_location_candidates(out)


# ------------------------------------------------------------
# Fallback column choice helpers
# ------------------------------------------------------------
def _best_fallback_name_col(columns: List[str]) -> Optional[str]:
    if not columns:
        return None

    normalized = [(c, _normalize_header(c)) for c in columns]

    for col, norm in normalized:
        if "name" in norm:
            return col

    for col, norm in normalized:
        if any(tok in norm for tok in ["product", "store", "company", "brand", "segment", "audience"]):
            return col

    return columns[0]


def _best_location_name_col(columns: List[str]) -> Optional[str]:
    normalized = [(c, _normalize_header(c)) for c in columns]
    for col, norm in normalized:
        if norm in {"store", "store name", "location", "location name", "branch", "branch name"}:
            return col
    return _best_fallback_name_col(columns)


def _assemble_address_from_row(row: Dict[str, str], address_cols: Dict[str, Optional[str]]) -> str:
    parts: List[str] = []

    for key in ["address", "address2", "city", "state", "zip", "country"]:
        col = address_cols.get(key)
        if col:
            val = _clean_text(row.get(col, ""))
            if val:
                parts.append(val)

    if parts:
        return ", ".join(parts)

    for col, val in row.items():
        norm = _normalize_header(col)
        if "address" in norm and _clean_text(val):
            return _clean_text(val)

    return ""


def _looks_like_entity_name(name: str) -> bool:
    name = _clean_text(name)
    if not name:
        return False
    if len(name) < 2:
        return False
    if len(name) > 200:
        return False
    if re.fullmatch(r"[\d\W_]+", name):
        return False
    return True


def _looks_like_location_candidate(name: str, address_raw: str) -> bool:
    if _clean_text(address_raw):
        return True
    return _looks_like_entity_name(name)


# ------------------------------------------------------------
# Deterministic merge
# ------------------------------------------------------------
def merge_product_candidates(items: List[ProductCandidate]) -> List[ProductCandidate]:
    merged: Dict[str, ProductCandidate] = {}

    for item in items:
        key = _normalize_key(item.name)
        if not key:
            continue

        prev = merged.get(key)
        if not prev:
            merged[key] = item
            continue

        best_desc = _choose_longer(prev.description_source, item.description_source)
        merged[key] = ProductCandidate(
            name=prev.name if len(prev.name) >= len(item.name) else item.name,
            description_source=best_desc,
            source=_merge_candidate_sources(prev.source, item.source),
        )

    return list(merged.values())


def merge_audience_candidates(items: List[AudienceCandidate]) -> List[AudienceCandidate]:
    merged: Dict[str, AudienceCandidate] = {}

    for item in items:
        key = _normalize_key(item.name)
        if not key:
            continue

        prev = merged.get(key)
        if not prev:
            merged[key] = item
            continue

        best_desc = _choose_longer(prev.description_source, item.description_source)
        merged[key] = AudienceCandidate(
            name=prev.name if len(prev.name) >= len(item.name) else item.name,
            description_source=best_desc,
            source=_merge_candidate_sources(prev.source, item.source),
        )

    return list(merged.values())


def merge_competitor_candidates(items: List[CompetitorCandidate]) -> List[CompetitorCandidate]:
    merged: Dict[str, CompetitorCandidate] = {}

    for item in items:
        name_key = _normalize_key(item.name)
        url_key = _normalize_key(item.url or "")
        key = url_key or name_key
        if not key:
            continue

        prev = merged.get(key)
        if not prev:
            merged[key] = item
            continue

        merged[key] = CompetitorCandidate(
            name=prev.name if len(prev.name) >= len(item.name) else item.name,
            url=prev.url or item.url,
            type_hint=prev.type_hint or item.type_hint,
            source=_merge_candidate_sources(prev.source, item.source),
        )

    return list(merged.values())


def merge_location_candidates(items: List[LocationCandidate]) -> List[LocationCandidate]:
    merged: Dict[str, LocationCandidate] = {}

    for item in items:
        key = f"{_normalize_key(item.name)}|{_normalize_key(item.address_raw or '')}"
        if key == "|":
            continue

        prev = merged.get(key)
        if not prev:
            merged[key] = item
            continue

        merged[key] = LocationCandidate(
            name=prev.name if len(prev.name) >= len(item.name) else item.name,
            address_raw=prev.address_raw or item.address_raw,
            phone=prev.phone or item.phone,
            landingPage=prev.landingPage or item.landingPage,
            timezone=prev.timezone or item.timezone,
            tags=_merge_string_lists(prev.tags, item.tags),
            hours_raw=prev.hours_raw or item.hours_raw,
            targeting_raw=prev.targeting_raw or item.targeting_raw,
            source=_merge_candidate_sources(prev.source, item.source),
        )

    return list(merged.values())


def _choose_longer(a: Optional[str], b: Optional[str]) -> Optional[str]:
    a = _clean_text(a)
    b = _clean_text(b)
    return a if len(a) >= len(b) else b


def _merge_string_lists(a: Optional[List[str]], b: Optional[List[str]]) -> Optional[List[str]]:
    vals = []
    for part in (a or []) + (b or []):
        cleaned = _clean_text(part)
        if cleaned and cleaned not in vals:
            vals.append(cleaned)
    return vals or None


def _merge_candidate_sources(a: CandidateSource, b: CandidateSource) -> CandidateSource:
    merged = dict(a.unused_columns)
    for key, value in b.unused_columns.items():
        if key not in merged and _clean_text(value):
            merged[key] = value
    return CandidateSource(unused_columns=merged)


# ------------------------------------------------------------
# Finalization Helpers
# ------------------------------------------------------------
def _chunk_list(items: List[Any], size: int) -> List[List[Any]]:
    return [items[i:i + size] for i in range(0, len(items), size)]


def _suggest_parallelism(model: str) -> int:
    model_l = (model or "").lower()
    if "nano" in model_l:
        return 4
    if "mini" in model_l:
        return 4
    return 3


def _suggest_max_output_tokens(doc_type: DocType, batch_size: int) -> int:
    if doc_type == "locations-doc":
        return 12000 if batch_size <= 30 else 16000
    if doc_type == "competitors-doc":
        return 8000
    if doc_type in {"products-doc", "audiences-doc"}:
        return 8000
    return 8000


def _normalize_result_key(v: str) -> str:
    s = (v or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s.strip(" -_|/.,;")


def _pick_richer_location_result(a: Dict[str, Any], b: Dict[str, Any]) -> Dict[str, Any]:
    def score(x: Dict[str, Any]) -> int:
        total = 0
        if x.get("name"):
            total += 2

        address = x.get("address") or {}
        if address.get("place_name"):
            total += 3
        if address.get("place_name_parts"):
            total += 2

        for field in ["timezone", "tags", "targeting", "phone", "landingPage", "hours"]:
            value = x.get(field)
            if value not in (None, "", [], {}):
                total += 1

        return total

    return a if score(a) >= score(b) else b


def _dedupe_location_results(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    merged: Dict[str, Dict[str, Any]] = {}

    for item in items:
        name = item.get("name") or ""
        place_name = ((item.get("address") or {}).get("place_name")) or ""
        key = f"{_normalize_result_key(name)}|{_normalize_result_key(place_name)}"

        if key == "|":
            continue

        prev = merged.get(key)
        if not prev:
            merged[key] = item
            continue

        merged[key] = _pick_richer_location_result(prev, item)

    return list(merged.values())


def _pick_richer_named_result(a: Dict[str, Any], b: Dict[str, Any], extra_fields: List[str]) -> Dict[str, Any]:
    def score(x: Dict[str, Any]) -> int:
        total = 0
        if x.get("name"):
            total += 2
        for field in extra_fields:
            value = x.get(field)
            if value not in (None, "", [], {}):
                total += 1
                if isinstance(value, str):
                    total += min(len(value) // 40, 3)
        return total

    return a if score(a) >= score(b) else b


def _dedupe_product_results(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    merged: Dict[str, Dict[str, Any]] = {}

    for item in items:
        key = _normalize_result_key(item.get("name") or "")
        if not key:
            continue

        prev = merged.get(key)
        if not prev:
            merged[key] = item
            continue

        merged[key] = _pick_richer_named_result(prev, item, ["description"])

    return list(merged.values())


def _dedupe_audience_results(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    merged: Dict[str, Dict[str, Any]] = {}

    for item in items:
        key = _normalize_result_key(item.get("name") or "")
        if not key:
            continue

        prev = merged.get(key)
        if not prev:
            merged[key] = item
            continue

        merged[key] = _pick_richer_named_result(prev, item, ["description"])

    return list(merged.values())


def _dedupe_competitor_results(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    merged: Dict[str, Dict[str, Any]] = {}

    for item in items:
        name_key = _normalize_result_key(item.get("name") or "")
        url_key = _normalize_result_key(item.get("url") or "")
        key = url_key or name_key
        if not key:
            continue

        prev = merged.get(key)
        if not prev:
            merged[key] = item
            continue

        merged[key] = _pick_richer_named_result(prev, item, ["url", "type"])

    return list(merged.values())


def _dedupe_final_results(doc_type: DocType, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if doc_type == "locations-doc":
        return _dedupe_location_results(items)
    if doc_type == "products-doc":
        return _dedupe_product_results(items)
    if doc_type == "audiences-doc":
        return _dedupe_audience_results(items)
    if doc_type == "competitors-doc":
        return _dedupe_competitor_results(items)
    return items


# ------------------------------------------------------------
# OpenAI finalization
# ------------------------------------------------------------
def call_structured_response(
    client: OpenAI,
    model: str,
    system_prompt: str,
    user_prompt: str,
    schema: Dict[str, Any],
    max_output_tokens: int = 12000,
) -> Dict[str, Any]:
    response = client.responses.create(
        model=model,
        instructions=system_prompt,
        input=user_prompt,
        max_output_tokens=max_output_tokens,
        text={
            "format": {
                "type": "json_schema",
                "name": "brand_doc_analysis_report",
                "strict": False,
                "schema": schema,
            }
        },
    )

    if not getattr(response, "output_text", None):
        raise ValueError("Empty model output")

    return json.loads(response.output_text)


def _drop_empty_values(d: Dict[str, Any]) -> Dict[str, Any]:
    return {
        k: v
        for k, v in d.items()
        if v not in (None, "", [], {}, ())
    }


def _candidate_to_payload(candidate: BaseModel) -> Dict[str, Any]:
    if isinstance(candidate, ProductCandidate):
        payload = _drop_empty_values({
            "name": candidate.name,
            "description_source": candidate.description_source,
        })
        if candidate.source.unused_columns:
            payload["_source"] = {"unused_columns": candidate.source.unused_columns}
        return payload

    if isinstance(candidate, AudienceCandidate):
        payload = _drop_empty_values({
            "name": candidate.name,
            "description_source": candidate.description_source,
        })
        if candidate.source.unused_columns:
            payload["_source"] = {"unused_columns": candidate.source.unused_columns}
        return payload

    if isinstance(candidate, CompetitorCandidate):
        payload = _drop_empty_values({
            "name": candidate.name,
            "url": candidate.url,
            "type_hint": candidate.type_hint,
        })
        if candidate.source.unused_columns:
            payload["_source"] = {"unused_columns": candidate.source.unused_columns}
        return payload

    if isinstance(candidate, LocationCandidate):
        payload = _drop_empty_values({
            "name": candidate.name,
            "address_raw": candidate.address_raw,
            "phone": candidate.phone,
            "landingPage": candidate.landingPage,
            "timezone": candidate.timezone,
            "tags": candidate.tags,
            "hours_raw": candidate.hours_raw,
            "targeting_raw": candidate.targeting_raw,
        })
        if candidate.source.unused_columns:
            payload["_source"] = {"unused_columns": candidate.source.unused_columns}
        return payload

    return candidate.model_dump()


def _build_candidates_payload(
    doc_type: DocType,
    candidates: List[BaseModel],
    parsed: ParsedTabularFile,
) -> Dict[str, Any]:
    trimmed = candidates[:MAX_FINALIZE_CANDIDATES]

    payload = {
        "docType": doc_type,
        "fileShape": {
            "sheetCount": len(parsed.sheets),
            "sheets": [
                {
                    "title": s.title,
                    "columns": s.columns,
                    "rowCount": len(s.rows),
                }
                for s in parsed.sheets
            ],
        },
        "candidateCount": len(trimmed),
        "candidates": [_candidate_to_payload(c) for c in trimmed],
    }

    encoded = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    if len(encoded) <= MAX_LLM_INPUT_CHARS:
        return payload

    compressed: List[Dict[str, Any]] = []
    current_size = 0

    for c in trimmed:
        item = _candidate_to_payload(c)
        item_json = json.dumps(item, ensure_ascii=False, separators=(",", ":"))
        if current_size + len(item_json) > MAX_LLM_INPUT_CHARS - 8000:
            break
        compressed.append(item)
        current_size += len(item_json)

    return {
        "docType": doc_type,
        "fileShape": {
            "sheetCount": len(parsed.sheets),
            "sheets": [
                {
                    "title": s.title,
                    "columns": s.columns,
                    "rowCount": len(s.rows),
                }
                for s in parsed.sheets
            ],
        },
        "candidateCount": len(compressed),
        "candidates": compressed,
        "truncated": True,
    }


async def finalize_with_llm(
    client: OpenAI,
    model: str,
    doc_type: DocType,
    brand_prompt: str,
    schema: Dict[str, Any],
    candidates: List[BaseModel],
    parsed: ParsedTabularFile,
) -> Dict[str, Any]:
    """
    Parallel batched finalization for all doc types when candidate count > 30.
    - <= 30 candidates: one call
    - > 30 candidates: parallel batches of max 30 candidates
    """
    system_prompt = get_finalize_system_prompt(doc_type)
    candidate_count = len(candidates)
    batch_size = 30

    if candidate_count <= batch_size:
        payload = _build_candidates_payload(doc_type, candidates, parsed)
        user_prompt = build_finalize_user_prompt(doc_type, brand_prompt, payload)
        max_output_tokens = _suggest_max_output_tokens(doc_type, candidate_count)

        logger.info(
            f"Finalizing {candidate_count} candidates with one LLM call for {doc_type} "
            f"(model={model}, max_output_tokens={max_output_tokens})"
        )

        return await asyncio.to_thread(
            call_structured_response,
            client,
            model,
            system_prompt,
            user_prompt,
            schema,
            max_output_tokens,
        )

    batches = _chunk_list(candidates, batch_size)
    parallelism = _suggest_parallelism(model)
    semaphore = asyncio.Semaphore(parallelism)

    logger.info(
        f"Finalizing {candidate_count} candidates in {len(batches)} parallel batches for {doc_type} "
        f"(model={model}, batch_size={batch_size}, parallelism={parallelism})"
    )

    async def _run_batch(batch_idx: int, batch_candidates: List[BaseModel]) -> Dict[str, Any]:
        async with semaphore:
            payload = _build_candidates_payload(doc_type, batch_candidates, parsed)
            user_prompt = build_finalize_user_prompt(doc_type, brand_prompt, payload)
            max_output_tokens = _suggest_max_output_tokens(doc_type, len(batch_candidates))

            logger.info(
                f"Starting batch {batch_idx + 1}/{len(batches)} "
                f"for {doc_type} with {len(batch_candidates)} candidates "
                f"(max_output_tokens={max_output_tokens})"
            )

            result = await asyncio.to_thread(
                call_structured_response,
                client,
                model,
                system_prompt,
                user_prompt,
                schema,
                max_output_tokens,
            )

            result_count = len(result.get("analysisResults", []))
            logger.info(
                f"Finished batch {batch_idx + 1}/{len(batches)} for {doc_type} "
                f"-> {result_count} results"
            )
            return result

    batch_results = await asyncio.gather(
        *[_run_batch(i, batch) for i, batch in enumerate(batches)],
        return_exceptions=True,
    )

    merged_results: List[Dict[str, Any]] = []
    failures: List[str] = []

    for i, batch_result in enumerate(batch_results):
        if isinstance(batch_result, Exception):
            failures.append(f"batch {i + 1}: {batch_result}")
            logger.error(f"Batch {i + 1} failed for {doc_type}: {batch_result}")
            continue

        merged_results.extend(batch_result.get("analysisResults", []))

    if not merged_results:
        raise ValueError(
            f"All finalization batches failed for {doc_type}"
            + (f" ({'; '.join(failures)})" if failures else "")
        )

    merged_results = _dedupe_final_results(doc_type, merged_results)

    logger.info(
        f"Batched finalization complete for {doc_type}: "
        f"{len(merged_results)} merged results from {candidate_count} input candidates"
    )

    return {"analysisResults": merged_results}


def deterministic_finalize(doc_type: DocType, candidates: List[BaseModel]) -> Dict[str, Any]:
    if doc_type == "products-doc":
        return {
            "analysisResults": [
                ProductItem(name=c.name, description=c.description_source or c.name).model_dump()
                for c in candidates
                if isinstance(c, ProductCandidate) and c.name
            ]
        }

    if doc_type == "audiences-doc":
        return {
            "analysisResults": [
                AudienceItem(name=c.name, description=c.description_source or c.name).model_dump()
                for c in candidates
                if isinstance(c, AudienceCandidate) and c.name
            ]
        }

    if doc_type == "locations-doc":
        results: List[Dict[str, Any]] = []
        for c in candidates:
            if not isinstance(c, LocationCandidate):
                continue
            if not c.name or not c.address_raw:
                continue

            address = AddressObj(
                place_name=c.address_raw,
                place_name_parts=_find_part_offsets(c.address_raw),
            )

            item = LocationItem(
                name=c.name,
                address=address,
                timezone=c.timezone or None,
                tags=c.tags or None,
                targeting=None,
                phone=c.phone or None,
                landingPage=c.landingPage or None,
                hours=None,
            )
            results.append(item.model_dump())

        return {"analysisResults": results}

    raise ValueError(f"Deterministic finalization not supported for doc_type={doc_type}")


# ------------------------------------------------------------
# Validation
# ------------------------------------------------------------
def validate_final_result(doc_type: DocType, result: Dict[str, Any]) -> Dict[str, Any]:
    model_cls = RESPONSE_MODEL_BY_DOC_TYPE[doc_type]
    validated = model_cls.model_validate(result)
    return validated.model_dump()


# ------------------------------------------------------------
# Endpoint
# ------------------------------------------------------------
@router.post("/brand-doc-analysis")
async def analyze_brand_document_tabular_fast(
    file: UploadFile = File(...),
    filePath: str = Form(...),
    docType: DocType = Form(...),
    brandId: str = Form(...),
    useAi: bool = Form(...),
    payload: Dict[str, Any] = Depends(verify_jwt),
) -> Dict[str, Any]:
    """
    Fast CSV/XLSX hybrid extractor.

    Flow:
    1. Validate/auth
    2. Parse CSV/XLSX into structured sheets
    3. Build candidates deterministically in Python
    4. Attach _source.unused_columns to each candidate
    5. Merge/dedupe deterministically in Python
    6. Always do exactly one final OpenAI structured call
    7. Validate strict final schema
    """
    user_id = await _get_verified_user(payload)
    _verify_brand_access(user_id, brandId)

    try:
        extension = detect_extension(file.filename or filePath, file.content_type)
        file_bytes = await file.read()

        if not file_bytes:
            raise HTTPException(status_code=400, detail="Uploaded file is empty")

        if len(file_bytes) > MAX_UPLOAD_BYTES:
            raise HTTPException(
                status_code=400,
                detail=f"Uploaded file exceeds max size of {MAX_UPLOAD_BYTES} bytes",
            )

        parsed = parse_tabular_file(file_bytes, extension)

        total_rows = sum(len(s.rows) for s in parsed.sheets)
        total_sheets = len(parsed.sheets)

        if total_rows == 0:
            raise HTTPException(status_code=400, detail="No parsable tabular content found")

        brand_config = get_brand_config(brandId)
        brand_prompt = get_brand_prompt(brand_config)

        candidates = build_candidates(parsed, docType)
        if not candidates:
            return {
                "analysisResults": [],
                "meta": {
                    "route": "tabular-fast",
                    "llmFinalizationUsed": False,
                    "sheetsParsed": total_sheets,
                    "rowsParsed": total_rows,
                    "candidatesBuilt": 0,
                },
            }
        
        if not useAi:
          final_result = deterministic_finalize(docType, candidates)
          validated = validate_final_result(docType, final_result)
          return {
              **validated,
              "meta": {
                  "route": "tabular-fast",
                  "llmFinalizationUsed": False,
                  "sheetsParsed": total_sheets,
                  "rowsParsed": total_rows,
                  "candidatesBuilt": len(candidates),
              },
          }

        client = get_openai_client()
        schema = get_schema(docType)
        final_result = await finalize_with_llm(
            client=client,
            model=MODEL_NAME,
            doc_type=docType,
            brand_prompt=brand_prompt,
            schema=schema,
            candidates=candidates,
            parsed=parsed,
        )

        validated = validate_final_result(docType, final_result)

        return {
            **validated,
            "meta": {
                "route": "tabular-fast",
                "llmFinalizationUsed": True,
                "sheetsParsed": total_sheets,
                "rowsParsed": total_rows,
                "candidatesBuilt": len(candidates),
            },
        }

    except ValidationError as e:
        logger.exception(f"Output validation failed: {e}")
        raise HTTPException(
            status_code=500,
            detail="Model returned invalid structured output",
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Tabular fast analysis failed: {e}")
        raise HTTPException(
            status_code=500,
            detail=str(e) or "Failed to analyze uploaded document",
        )